import openpyxl
import re
import pyautogui
import time

def extract_dates_dollar_amounts_account_numbers(input_file, output_file):
    # Load the input workbook and worksheets
    input_wb = openpyxl.load_workbook(input_file)
    ws_input = input_wb["Sheet1"]  # Input worksheet
    ws_output = input_wb["NumAcc Bot"]  # Output worksheet

    # Define input range and output starting cells
    input_range = ws_input["A1:A10"]  # Adjust range as needed
    date_cell = ws_output["E32"]  # Starting cell for dates
    dollar_cell = ws_output["G32"]  # Starting cell for dollar amounts
    account_cell = ws_output["H32"]  # Starting cell for account numbers

    # Regular expressions for patterns
    date_pattern_without_year = r"\b(January|February|March|April|May|June|July|August|September|October|November|December) \d{1,2}\b"
    date_pattern_with_typo = r"\b(January|February|March|April|May|June|July|August|September|October|November|December) \d{1,2}, \d{2}\b"
    date_pattern_valid = r"\b(January|February|March|April|May|June|July|August|September|October|November|December) \d{1,2}, \d{4}\b"
    dollar_pattern = r"\$\d+(,\d{3})*(\.\d{2})?"
    account_pattern = r"\b\d{4}\b"

    # Track already processed items
    already_processed = set()

    # Loop through each cell in the range
    for cell in input_range:
        cell_value = cell.value
        if cell_value:
            # Extract valid dates first
            valid_matches = re.findall(date_pattern_valid, cell_value)
            for match in valid_matches:
                if match not in already_processed:
                    date_cell.value = match
                    date_cell.offset(row=0, column=1).value = "Valid Date"
                    already_processed.add(match)
                    date_cell = ws_output.cell(row=date_cell.row + 1, column=date_cell.column)

            # Extract dates with typos in years
            typo_matches = re.findall(date_pattern_with_typo, cell_value)
            for match in typo_matches:
                if match not in already_processed:
                    date_cell.value = match
                    date_cell.offset(row=0, column=1).value = "Year Typo Found"
                    already_processed.add(match)
                    date_cell = ws_output.cell(row=date_cell.row + 1, column=date_cell.column)

            # Extract dates without years
            no_year_matches = re.findall(date_pattern_without_year, cell_value)
            for match in no_year_matches:
                if match not in already_processed:
                    date_cell.value = match
                    date_cell.offset(row=0, column=1).value = "No Year Found"
                    already_processed.add(match)
                    date_cell = ws_output.cell(row=date_cell.row + 1, column=date_cell.column)

            # Extract dollar amounts
            dollar_matches = re.findall(dollar_pattern, cell_value)
            for match in dollar_matches:
                if match not in already_processed:
                    dollar_cell.value = match
                    already_processed.add(match)
                    dollar_cell = ws_output.cell(row=dollar_cell.row + 1, column=dollar_cell.column)

            # Extract account numbers (exactly 4 digits)
            account_matches = re.findall(account_pattern, cell_value)
            for match in account_matches:
                if match not in already_processed:
                    account_cell.value = match
                    already_processed.add(match)
                    account_cell = ws_output.cell(row=account_cell.row + 1, column=account_cell.column)

    # Save the workbook
    input_wb.save(output_file)
    print("Extraction complete: Dates, dollar amounts, and account numbers!")

def clear_and_paste_into_one_cell(input_file, target_sheet, target_cell):
    # Load the workbook and target worksheet
    wb = openpyxl.load_workbook(input_file)
    ws_target = wb[target_sheet]

    # Clear contents of the target sheet
    for row in ws_target.iter_rows():
        for cell in row:
            cell.value = None

    # Save the workbook after clearing the target sheet
    wb.save(input_file)

    # Activate the target cell in edit mode (simulate user actions)
    # NOTE: This requires Excel to be open and the correct cell selected manually.
    time.sleep(1)  # Allow time for manual selection (adjust as needed)
    pyautogui.press('f2')  # Simulate pressing F2 to activate edit mode
    pyautogui.hotkey('ctrl', 'v')  # Simulate pressing Ctrl + V to paste

    print("Paste operation simulated.")

# Example usage
# Step 1: Extract data
extract_dates_dollar_amounts_account_numbers("input.xlsx", "output.xlsx")

# Step 2: Clear and paste into one cell
clear_and_paste_into_one_cell("output.xlsx", "TargetSheet", "A1")
